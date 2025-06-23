import numpy as np
import openpyxl
import xlrd
from scipy.interpolate import interp1d, make_interp_spline

from odym.functions.utils import Tuple_MI


def ReadParameter(
    ParPath,
    ThisPar,
    ThisParIx,
    IndexMatch,
    ThisParLayerSel,
    MasterClassification,
    IndexTable,
    IndexTable_ClassificationNames,
    ScriptConfig,
    Mylog,
):
    """
    This function reads a model parameter from the corresponding parameter file
    """
    Parfile = xlrd.open_workbook(ParPath + ".xlsx")
    ParHeader = Parfile.sheet_by_name("Cover")

    IM = eval(IndexMatch)  # List that matches model aspects to parameter indices

    ri = 1  # row index
    MetaData = {}
    while True:  # read cover sheet info
        ThisItem = ParHeader.cell_value(ri, 0)
        if ThisItem != "Dataset_RecordType":
            MetaData[ThisItem] = ParHeader.cell_value(ri, 1)
            ri += 1
        else:
            break  # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType

    # Check whether parameter file uses same classification:
    if (
        "ODYM_Classifications_Master_"
        + ScriptConfig["Version of master classification"]
        != MetaData["Dataset_Classification_version_number"]
    ):
        Mylog.critical(
            "CLASSIFICATION FILE FATAL ERROR: Classification file of parameter "
            + ThisPar
            + " is not identical to the classification master file used for the current model run."
        )

    if ParHeader.cell_value(ri, 1) == "List":
        IList = []
        IListMeaning = []
        ci = 1  # column index
        while True:
            if ParHeader.cell_value(ri + 1, ci) != "":
                IList.append(ParHeader.cell_value(ri + 1, ci))
                IListMeaning.append(ParHeader.cell_value(ri + 2, ci))
                ci += 1
            else:
                break
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]

        ValueList = []
        VIComment = []
        ci = 1  # column index
        while True:
            if ParHeader.cell_value(ri + 4, ci) != "":
                ValueList.append(ParHeader.cell_value(ri + 3, ci))
                VIComment.append(ParHeader.cell_value(ri + 4, ci))
                ci += 1
            else:
                break

        # Check whether all indices are present in the index table of the model
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != IList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + IList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )

        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name("Values_Master")
        ColOffset = len(IList)
        RowOffset = 1  # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx = 0
        while True:
            try:
                CV = ValuesSheet.cell_value(cx + RowOffset, ColOffset)
            except:
                break
            TargetPosition = []
            for mx in range(
                0, len(IList)
            ):  # mx iterates over the aspects of the parameter
                CurrentItem = ValuesSheet.cell_value(cx + RowOffset, IM[mx])
                try:
                    TargetPosition.append(
                        IndexTable.set_index("IndexLetter")
                        .loc[ThisParIx[mx]]
                        .Classification.Items.index(CurrentItem)
                    )
                except:
                    break  # Current parameter value is not needed for model, outside scope for a certain aspect.
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
            cx += 1

        Mylog.info(
            "A total of "
            + str(cx + 1)
            + " values was read from file for parameter "
            + ThisPar
            + "."
        )
        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )

    ### Table version ###
    if (
        ParHeader.cell_value(ri, 1) == "Table"
    ):  # have 3 while loops, one for row indices, one for column indices, one for value layers

        RIList = []
        RISize = []
        RIListMeaning = []
        ci = 1  # column index
        while True:
            if ParHeader.cell_value(ri + 1, ci) != "":
                RIList.append(ParHeader.cell_value(ri + 1, ci))
                RISize.append(int(ParHeader.cell_value(ri + 2, 1)))
                RIListMeaning.append(ParHeader.cell_value(ri + 3, ci))
                ci += 1
            else:
                break
        RISize = RISize[0]

        CIList = []
        CISize = []
        CIListMeaning = []
        ci = 1  # column index
        while True:
            if ParHeader.cell_value(ri + 4, ci) != "":
                CIList.append(ParHeader.cell_value(ri + 4, ci))
                CISize.append(int(ParHeader.cell_value(ri + 5, 1)))
                CIListMeaning.append(ParHeader.cell_value(ri + 6, ci))
                ci += 1
            else:
                break
        CISize = CISize[0]

        # Re-Order indices to fit model aspect order:
        ComIList = RIList + CIList
        ComIList = [ComIList[i] for i in IM]

        ValueList = []
        VIComment = []
        ci = 1  # column index
        while True:
            if ParHeader.cell_value(ri + 7, ci) != "":
                ValueList.append(ParHeader.cell_value(ri + 7, ci))
                VIComment.append(ParHeader.cell_value(ri + 8, ci))
                ci += 1
            else:
                break

        # Check whether all indices are present in the index table of the model
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Row index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Column index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0, len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])
        CIIndexLetter = []
        for m in range(0, len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m + len(RIList))])

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + ComIList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )

        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name(ValueList[ThisParLayerSel[0]])
        ColOffset = len(RIList)
        RowOffset = len(CIList)
        RowNos = RISize
        ColNos = CISize

        TargetPos_R = []
        for m in range(0, RowNos):
            TP_RD = []
            for mc in range(0, len(RIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(m + RowOffset, mc))
                except:
                    CurrentItem = ValuesSheet.cell_value(m + RowOffset, mc)
                try:
                    IX = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[RIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_RD.append((IX, TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)

        TargetPos_C = []
        for n in range(0, ColNos):
            TP_CD = []
            for mc in range(0, len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(mc, n + ColOffset))
                except:
                    CurrentItem = ValuesSheet.cell_value(mc, n + ColOffset)
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[CIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_CD.append((IX, TPIX))
                except:
                    TP_CD.append(None)
                    break
            TargetPos_C.append(TP_CD)

        for m in range(0, RowNos):
            for n in range(0, ColNos):
                TargetPosition = [0 for i in range(0, len(ComIList))]
                try:
                    for i in range(0, len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1]
                    for i in range(0, len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1]
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(ComIList):
                    Values[tuple(TargetPosition)] = ValuesSheet.cell_value(
                        m + RowOffset, n + ColOffset
                    )
                    ValIns[tuple(TargetPosition)] = 1

        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )

    return MetaData, Values


def ReadParameterV2(
    ParPath,
    ThisPar,
    ThisParIx,
    IndexMatch,
    ThisParLayerSel,
    MasterClassification,
    IndexTable,
    IndexTable_ClassificationNames,
    ScriptConfig,
    Mylog,
    ParseUncertainty,
):
    """
    This function reads a model parameter from the corresponding parameter file
    """
    Parfile = xlrd.open_workbook(ParPath + ".xlsx")
    ParHeader = Parfile.sheet_by_name("Cover")

    IM = eval(IndexMatch)  # List that matches model aspects to parameter indices

    ri = 1  # row index
    MetaData = {}
    while True:  # read cover sheet info
        ThisItem = ParHeader.cell_value(ri, 0)
        if ThisItem != "[Empty on purpose]" and ThisItem != "Dataset_RecordType":
            MetaData[ThisItem] = ParHeader.cell_value(ri, 1)
            if ThisItem == "Dataset_Unit":
                if ParHeader.cell_value(ri, 1) == "GLOBAL":
                    MetaData["Unit_Global"] = ParHeader.cell_value(ri, 2)
                    MetaData["Unit_Global_Comment"] = ParHeader.cell_value(ri, 3)
            if ThisItem == "Dataset_Uncertainty":
                # if LIST is specified, nothing happens here.
                if ParHeader.cell_value(ri, 1) == "GLOBAL":
                    MetaData["Dataset_Uncertainty_Global"] = ParHeader.cell_value(ri, 2)
                if ParHeader.cell_value(ri, 1) == "TABLE":
                    MetaData["Dataset_Uncertainty_Sheet"] = ParHeader.cell_value(ri, 2)
            if ThisItem == "Dataset_Comment":
                if ParHeader.cell_value(ri, 1) == "GLOBAL":
                    MetaData["Dataset_Comment_Global"] = ParHeader.cell_value(ri, 2)
            ri += 1
        else:
            break  # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType

    # Check whether parameter file uses same classification:
    if (
        ScriptConfig["Version of master classification"]
        != MetaData["Dataset_Classification_version_number"]
    ):
        Mylog.critical(
            "CLASSIFICATION FILE FATAL ERROR: Classification file of parameter "
            + ThisPar
            + " is not identical to the classification master file used for the current model run."
        )

    # Continue parsing until line 'Dataset_RecordType' is found:
    while True:
        ThisItem = ParHeader.cell_value(ri, 0)
        if ThisItem == "Dataset_RecordType":
            break
        else:
            ri += 1

    ### List version ###
    if ParHeader.cell_value(ri, 1) == "LIST":
        IList = []
        IListMeaning = []
        RI_Start = ri + 2
        while True:
            if ParHeader.cell_value(RI_Start, 0) != "":
                IList.append(ParHeader.cell_value(RI_Start, 0))
                IListMeaning.append(ParHeader.cell_value(RI_Start, 1))
                RI_Start += 1
            else:
                break
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]

        ValueList = []
        VIComment = []
        RI_Start = ri + 2
        while True:
            if ParHeader.cell_value(RI_Start, 2) != "":
                ValueList.append(ParHeader.cell_value(RI_Start, 2))
                VIComment.append(ParHeader.cell_value(RI_Start, 3))
                RI_Start += 1
            else:
                break

        # Check whether all indices are present in the index table of the model
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != IList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + IList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )

        # Read parameter values into array, uncertainty into list:
        Values = np.zeros((IndexSizesM))  # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM)  # parameter value uncertainties
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name("Values_Master")
        ColOffset = len(IList)
        RowOffset = 1  # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx = 0
        while True:
            try:
                CV = ValuesSheet.cell_value(cx + RowOffset, ColOffset)
            except:
                break
            TargetPosition = []
            for mx in range(
                0, len(IList)
            ):  # mx iterates over the aspects of the parameter
                CurrentItem = ValuesSheet.cell_value(cx + RowOffset, IM[mx])
                try:
                    TargetPosition.append(
                        IndexTable.set_index("IndexLetter")
                        .loc[ThisParIx[mx]]
                        .Classification.Items.index(CurrentItem)
                    )
                except:
                    break  # Current parameter value is not needed for model, outside scope for a certain aspect.
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
                Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = (
                    ValuesSheet.cell_value(cx + RowOffset, ColOffset + 3)
                )
            cx += 1

        Mylog.info(
            "A total of "
            + str(cx)
            + " values was read from file for parameter "
            + ThisPar
            + "."
        )
        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )

    ### Table version ###
    if (
        ParHeader.cell_value(ri, 1) == "TABLE"
    ):  # have 3 while loops, one for row indices, one for column indices, one for value layers
        ColNos = int(ParHeader.cell_value(ri, 5))  # Number of columns in dataset
        RowNos = int(ParHeader.cell_value(ri, 3))  # Number of rows in dataset

        RI = ri + 2  # row where indices start
        RIList = []
        RIListMeaning = []
        while True:
            if ParHeader.cell_value(RI, 0) != "":
                RIList.append(ParHeader.cell_value(RI, 0))
                RIListMeaning.append(ParHeader.cell_value(RI, 1))
                RI += 1
            else:
                break

        RI = ri + 2  # row where indices start
        CIList = []
        CIListMeaning = []
        while True:
            if ParHeader.cell_value(RI, 2) != "":
                CIList.append(ParHeader.cell_value(RI, 2))
                CIListMeaning.append(ParHeader.cell_value(RI, 3))
                RI += 1
            else:
                break

        # Re-Order indices to fit model aspect order:
        ComIList = RIList + CIList  # List of all indices, both rows and columns
        ComIList = [ComIList[i] for i in IM]

        RI = ri + 2  # row where indices start
        ValueList = []
        VIComment = []
        while True:
            if ParHeader.cell_value(RI, 4) != "":
                ValueList.append(ParHeader.cell_value(RI, 4))
                VIComment.append(ParHeader.cell_value(RI, 5))
                RI += 1
            else:
                break

        # Check whether all indices are present in the index table of the model
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Row index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Column index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0, len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])
        CIIndexLetter = []
        for m in range(0, len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m + len(RIList))])

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + ComIList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )

        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))  # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM)  # parameter value uncertainties
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded, contains 0 or 1.
        ValuesSheet = Parfile.sheet_by_name(ValueList[ThisParLayerSel[0]])
        if ParseUncertainty:
            if "Dataset_Uncertainty_Sheet" in MetaData:
                UncertSheet = Parfile.sheet_by_name(
                    MetaData["Dataset_Uncertainty_Sheet"]
                )
        ColOffset = len(RIList)
        RowOffset = len(CIList)
        cx = 0

        TargetPos_R = []  # Determine all row target positions in data array
        for m in range(0, RowNos):
            TP_RD = []
            for mc in range(0, len(RIList)):
                try:
                    CurrentItem = int(
                        ValuesSheet.cell_value(m + RowOffset, mc)
                    )  # in case items come as int, e.g., years
                except:
                    CurrentItem = ValuesSheet.cell_value(m + RowOffset, mc)
                try:
                    IX = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[RIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_RD.append((IX, TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)

        TargetPos_C = []  # Determine all col target positions in data array
        for n in range(0, ColNos):
            TP_CD = []
            for mc in range(0, len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(mc, n + ColOffset))
                except:
                    CurrentItem = ValuesSheet.cell_value(mc, n + ColOffset)
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[CIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_CD.append((IX, TPIX))
                except:
                    TP_CD.append(None)
                    break
            TargetPos_C.append(TP_CD)

        for m in range(0, RowNos):  # Read values from excel template
            for n in range(0, ColNos):
                TargetPosition = [0 for i in range(0, len(ComIList))]
                try:
                    for i in range(0, len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1]
                    for i in range(0, len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1]
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(
                    ComIList
                ):  # Read value if TargetPosition Tuple has same length as indexList
                    Values[tuple(TargetPosition)] = ValuesSheet.cell_value(
                        m + RowOffset, n + ColOffset
                    )
                    ValIns[tuple(TargetPosition)] = 1
                    # Add uncertainty
                    if ParseUncertainty:
                        if "Dataset_Uncertainty_Global" in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = (
                                MetaData["Dataset_Uncertainty_Global"]
                            )
                        if "Dataset_Uncertainty_Sheet" in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = (
                                UncertSheet.cell_value(m + RowOffset, n + ColOffset)
                            )
                cx += 1

        Mylog.info(
            "A total of "
            + str(cx)
            + " values was read from file for parameter "
            + ThisPar
            + "."
        )
        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )
    if ParseUncertainty:
        return MetaData, Values, Uncertainty
    else:
        return MetaData, Values


def ReadParameterXLSX(
    ParPath,
    ThisPar,
    ThisParIx,
    IndexMatch,
    ThisParLayerSel,
    ThisParProcMethod,
    MasterClassification,
    IndexTable,
    IndexTable_ClassificationNames,
    ScriptConfig,
    Mylog,
    ParseUncertainty,
):
    """
    This function reads a model parameter from the corresponding parameter file and used openpyxl
    """
    Parfile = openpyxl.load_workbook(ParPath + ".xlsx", data_only=True)
    ParHeader = Parfile["Cover"]

    IM = eval(IndexMatch)  # List that matches model aspects to parameter indices

    ri = 2  # row index
    MetaData = {}
    while True:  # read cover sheet info
        ThisItem = ParHeader.cell(ri, 1).value
        if ThisItem != "[Empty on purpose]" and ThisItem != "Dataset_RecordType":
            MetaData[ThisItem] = ParHeader.cell(ri, 2).value
            if ThisItem == "Dataset_Unit":
                if ParHeader.cell(ri, 2).value == "GLOBAL":
                    MetaData["Unit_Global"] = ParHeader.cell(ri, 3).value
                    MetaData["Unit_Global_Comment"] = ParHeader.cell(ri, 4).value
            if ThisItem == "Dataset_Uncertainty":
                # if LIST is specified, nothing happens here.
                if ParHeader.cell(ri, 2).value == "GLOBAL":
                    MetaData["Dataset_Uncertainty_Global"] = ParHeader.cell(ri, 3).value
                if ParHeader.cell(ri, 2).value == "TABLE":
                    MetaData["Dataset_Uncertainty_Sheet"] = ParHeader.cell(ri, 3).value
            if ThisItem == "Dataset_Comment":
                if ParHeader.cell(ri, 2).value == "GLOBAL":
                    MetaData["Dataset_Comment_Global"] = ParHeader.cell(ri, 3).value
            ri += 1
        else:
            break  # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType

    # Check whether parameter file uses same classification:
    if (
        ScriptConfig["Version of master classification"]
        != MetaData["Dataset_Classification_version_number"]
    ):
        Mylog.critical(
            "CLASSIFICATION FILE FATAL ERROR: Classification file of parameter "
            + ThisPar
            + " is not identical to the classification master file used for the current model run."
        )

    # Continue parsing until line 'Dataset_RecordType' is found:
    while True:
        ThisItem = ParHeader.cell(ri, 1).value
        if ThisItem == "Dataset_RecordType":
            Mylog.info(ParHeader.cell(ri, 2).value)
            break
        else:
            ri += 1

    ### List version ###
    if ParHeader.cell(ri, 2).value == "LIST":  # ri = 21
        IList = []
        IListMeaning = []
        RI_Start = ri + 2
        while ParHeader.cell(RI_Start, 1).value is not None:
            IList.append(ParHeader.cell(RI_Start, 1).value)
            IListMeaning.append(ParHeader.cell(RI_Start, 2).value)
            RI_Start += 1
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]

        ValueList = []
        VIComment = []
        RI_Start = ri + 2
        while ParHeader.cell(RI_Start, 3).value is not None:
            ValueList.append(ParHeader.cell(RI_Start, 3).value)
            VIComment.append(ParHeader.cell(RI_Start, 4).value)
            RI_Start += 1

        # Check whether all indices are present in the index table of the model
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != IList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + IList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )
        # Read parameter values into array, uncertainty into list:
        Values = np.zeros((IndexSizesM))  # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM)  # parameter value uncertainties
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded
        ValuesSheet = Parfile["Values_Master"]
        ColOffset = len(IList)
        RowOffset = 1  # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx = 0
        while True:
            if ValuesSheet.cell(cx + RowOffset + 1, ColOffset + 1).value is not None:
                CV = ValuesSheet.cell(cx + RowOffset + 1, ColOffset + 1).value
            else:
                break
            TargetPosition = []
            for mx in range(
                0, len(IList)
            ):  # mx iterates over the aspects of the parameter
                CurrentItem = ValuesSheet.cell(cx + RowOffset + 1, IM[mx] + 1).value

                try:
                    TargetPosition.append(
                        IndexTable.set_index("IndexLetter")
                        .loc[ThisParIx[mx]]
                        .Classification.Items.index(CurrentItem)
                    )
                except:
                    break  # Current parameter value is not needed for model, outside scope for a certain aspect.
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
                Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = ValuesSheet.cell(
                    cx + RowOffset + 1, ColOffset + 4
                ).value
            cx += 1

        Mylog.info(
            "A total of "
            + str(cx)
            + " values was read from file for parameter "
            + ThisPar
            + "."
        )
        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )

    ### Table version ###
    if (
        ParHeader.cell(ri, 2).value == "TABLE"
    ):  # have 3 while loops, one for row indices, one for column indices, one for value layers
        ColNos = int(ParHeader.cell(ri, 6).value)  # Number of columns in dataset
        RowNos = int(ParHeader.cell(ri, 4).value)  # Number of rows in dataset

        RI = ri + 2  # row where indices start
        RIList = []
        RIListMeaning = []
        while True:
            if ParHeader.cell(RI, 1).value is not None:
                RIList.append(ParHeader.cell(RI, 1).value)
                RIListMeaning.append(ParHeader.cell(RI, 2).value)
                RI += 1
            else:
                break

        RI = ri + 2  # row where indices start
        CIList = []
        CIListMeaning = []
        while True:
            if ParHeader.cell(RI, 3).value is not None:
                CIList.append(ParHeader.cell(RI, 3).value)
                CIListMeaning.append(ParHeader.cell(RI, 4).value)
                RI += 1
            else:
                break

        # Re-Order indices to fit model aspect order:
        ComIList = RIList + CIList  # List of all indices, both rows and columns
        ComIList = [ComIList[i] for i in IM]

        RI = ri + 2  # row where indices start
        ValueList = []
        VIComment = []
        while True:
            if ParHeader.cell(RI, 5).value is not None:
                ValueList.append(ParHeader.cell(RI, 5).value)
                VIComment.append(ParHeader.cell(RI, 6).value)
                RI += 1
            else:
                break

        # Check whether all indices are present in the index table of the model
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Row index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error(
                "CLASSIFICATION ERROR: Column index list of data file for parameter "
                + ThisPar
                + " contains indices that are not part of the current model run."
            )

        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0, len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])
        CIIndexLetter = []
        for m in range(0, len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m + len(RIList))])

        # Check how well items match between model and data, select items to import
        IndexSizesM = []  # List of dimension size for model
        for m in range(0, len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName = (
                IndexTable.set_index("IndexLetter").loc[ThisDim].Classification.Name
            )
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error(
                    "CLASSIFICATION ERROR: Classification "
                    + ThisDimClassificationName
                    + " for aspect "
                    + ThisDim
                    + " of parameter "
                    + ThisPar
                    + " must be identical to the specified classification of the corresponding parameter dimension, which is "
                    + ComIList[m]
                )
                break  # Stop parsing parameter, will cause model to halt

            IndexSizesM.append(
                IndexTable.set_index("IndexLetter").loc[ThisDim]["IndexSize"]
            )

        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))  # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM)  # parameter value uncertainties
        ValIns = np.zeros(
            (IndexSizesM)
        )  # Array to check how many values are actually loaded, contains 0 or 1.
        ValuesSheet = Parfile[ValueList[ThisParLayerSel[0]]]
        if ParseUncertainty:
            if "Dataset_Uncertainty_Sheet" in MetaData:
                UncertSheet = Parfile[MetaData["Dataset_Uncertainty_Sheet"]]
        ColOffset = len(RIList)
        RowOffset = len(CIList)
        cx = 0

        TargetPos_R = []  # Determine all row target positions in data array
        for m in range(0, RowNos):
            TP_RD = []
            for mc in range(0, len(RIList)):
                try:
                    CurrentItem = int(
                        ValuesSheet.cell(m + RowOffset + 1, mc + 1).value
                    )  # in case items come as int, e.g., years
                except:
                    CurrentItem = ValuesSheet.cell(m + RowOffset + 1, mc + 1).value
                try:
                    IX = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[RIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_RD.append((IX, TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)

        TargetPos_C = []  # Determine all col target positions in data array
        for n in range(0, ColNos):
            TP_CD = []
            for mc in range(0, len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell(mc + 1, n + ColOffset + 1).value)
                except:
                    CurrentItem = ValuesSheet.cell(mc + 1, n + ColOffset + 1).value
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = (
                        IndexTable.set_index("IndexLetter")
                        .loc[CIIndexLetter[mc]]
                        .Classification.Items.index(CurrentItem)
                    )
                    TP_CD.append((IX, TPIX))
                except:
                    TP_CD.append(None)
                    break
            TargetPos_C.append(TP_CD)

        for m in range(0, RowNos):  # Read values from excel template
            for n in range(0, ColNos):
                TargetPosition = [0 for i in range(0, len(ComIList))]
                try:
                    for i in range(0, len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1]
                    for i in range(0, len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1]
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(
                    ComIList
                ):  # Read value if TargetPosition Tuple has same length as indexList
                    Values[tuple(TargetPosition)] = ValuesSheet.cell(
                        m + RowOffset + 1, n + ColOffset + 1
                    ).value
                    ValIns[tuple(TargetPosition)] = 1
                    # Add uncertainty
                    if ParseUncertainty:
                        if "Dataset_Uncertainty_Global" in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = (
                                MetaData["Dataset_Uncertainty_Global"]
                            )
                        if "Dataset_Uncertainty_Sheet" in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = (
                                UncertSheet.cell_value(
                                    m + RowOffset + 1, n + ColOffset + 1
                                )
                            )
                cx += 1

        Mylog.info(
            "A total of "
            + str(cx)
            + " values was read from file for parameter "
            + ThisPar
            + "."
        )
        Mylog.info(
            str(ValIns.sum())
            + " of "
            + str(np.prod(IndexSizesM))
            + " values for parameter "
            + ThisPar
            + " were assigned."
        )

    Processing_methods = eval(ThisParProcMethod)
    for processing in Processing_methods:

        if processing == "none":
            continue

        elif processing.startswith("replicate"):
            if len(ThisParProcMethod.split("_")) != 5:
                Mylog.error(
                    "Replicate processing error: instruction not recognized for parameter "
                    + ThisPar
                    + "."
                )

            replicateIndex = processing.split("_")[1]
            targetValue = processing.split("_")[2]
            copyValue = processing.split("_")[4]

            if replicateIndex not in ThisParIx:
                Mylog.error(
                    "Replicate processing error: index "
                    + replicateIndex
                    + " not a dimension for parameter "
                    + ThisPar
                    + "."
                )
            if (
                copyValue
                not in IndexTable.set_index("IndexLetter")
                .loc[replicateIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Replicate processing error: "
                    + copyValue
                    + " not in the classification for aspect "
                    + replicateIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )
            if (
                targetValue
                not in IndexTable.set_index("IndexLetter")
                .loc[replicateIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Replicate processing error: "
                    + targetValue
                    + " not in the classification for aspect "
                    + replicateIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )

            ix_position = ThisParIx.find(replicateIndex)
            C_ix = (
                IndexTable.set_index("IndexLetter")
                .loc[replicateIndex]
                .Classification.Items.index(copyValue)
            )
            T_ix = (
                IndexTable.set_index("IndexLetter")
                .loc[replicateIndex]
                .Classification.Items.index(targetValue)
            )
            dimensions = Values.shape
            for indices in np.ndindex(
                dimensions[:ix_position] + dimensions[ix_position + 1 :]
            ):
                Values[indices[:ix_position] + (T_ix,) + indices[ix_position:]] = (
                    Values[indices[:ix_position] + (C_ix,) + indices[ix_position:]]
                )
            Mylog.info(
                "Replicated "
                + copyValue
                + " values in "
                + targetValue
                + " for aspect "
                + replicateIndex
                + " for parameter "
                + ThisPar
                + "."
            )

        elif processing.startswith("interpolate"):
            if len(processing.split("_")) != 5:
                Mylog.error(
                    "Interpolate processing error: instruction not recognized for parameter "
                    + ThisPar
                    + "."
                )
            interpIndex = processing.split("_")[1]
            startValue = int(processing.split("_")[2])
            endValue = int(processing.split("_")[3])
            method = processing.split("_")[4]

            if interpIndex not in ThisParIx:
                Mylog.error(
                    "Interpolation processing error: index "
                    + interpIndex
                    + " not a dimension for parameter "
                    + ThisPar
                    + "."
                )
            if (
                startValue
                not in IndexTable.set_index("IndexLetter")
                .loc[interpIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Interpolation processing error: "
                    + str(startValue)
                    + " not in the classification for aspect "
                    + interpIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )
            if (
                endValue
                not in IndexTable.set_index("IndexLetter")
                .loc[interpIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Interpolation processing error: "
                    + str(endValue)
                    + " not in the classification for aspect "
                    + interpIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )

            startIndex = (
                IndexTable.set_index("IndexLetter")
                .loc[interpIndex]
                .Classification.Items.index(startValue)
            )
            endIndex = (
                IndexTable.set_index("IndexLetter")
                .loc[interpIndex]
                .Classification.Items.index(endValue)
            )
            ix_position = ThisParIx.find(interpIndex)
            ValIns_b = np.array(ValIns, dtype=bool)
            dimensions = Values.shape

            for indices in np.ndindex(
                dimensions[:ix_position] + dimensions[ix_position + 1 :]
            ):
                if (
                    ValIns_b[
                        indices[:ix_position]
                        + (startIndex,)
                        + indices[ix_position:]
                    ]
                    and ValIns_b[
                        indices[:ix_position] + (endIndex,) + indices[ix_position:]
                    ]
                ):
                    x = [
                        IndexTable.set_index("IndexLetter")
                        .loc[interpIndex]
                        .Classification.Items[m]
                        for m in range(startIndex, endIndex + 1)
                        if ValIns_b[
                            indices[:ix_position] + (m,) + indices[ix_position:]
                        ]
                    ]
                    y = [
                        Values[indices[:ix_position] + (m,) + indices[ix_position:]]
                        for m in range(startIndex, endIndex + 1)
                        if ValIns_b[
                            indices[:ix_position] + (m,) + indices[ix_position:]
                        ]
                    ]
                    if method == "spline":
                        clamped_spline = make_interp_spline(
                            x, y, bc_type=([(2, 0)], [(1, 0)])
                        )  # spline function, free (2nd derivative=0) for starting boundary condition and clamped (1st derivative=0) for end boundary condition
                        for m in range(startIndex, endIndex + 1):
                            Values[
                                indices[:ix_position] + (m,) + indices[ix_position:]
                            ] = clamped_spline(
                                IndexTable.set_index("IndexLetter")
                                .loc[interpIndex]
                                .Classification.Items[m]
                            )
                    elif method == "linear":
                        f = interp1d(x, y, kind="linear")
                        for m in range(startIndex, endIndex + 1):
                            Values[
                                indices[:ix_position] + (m,) + indices[ix_position:]
                            ] = f(
                                IndexTable.set_index("IndexLetter")
                                .loc[interpIndex]
                                .Classification.Items[m]
                            )
                    else:
                        Mylog.error(
                            "Interpolation error: method "
                            + method
                            + " not recognized for parameter "
                            + ThisPar
                            + "."
                        )
                        break

            Mylog.info(
                "Intrpolated "
                + str(interpIndex)
                + " aspect from "
                + str(startValue)
                + " to "
                + str(endValue)
                + " for parameter "
                + ThisPar
                + "."
            )
            count_neg = (Values < 0).sum()
            if count_neg > 0:
                Values[Values < 0] = 0
                Mylog.info(
                    str(count_neg)
                    + " negative values from spline interpolation set to 0."
                )

        elif processing.startswith("copy"):
            if len(processing.split("_")) != 5:
                Mylog.error(
                    "Copy processing error: instruction not recognized for parameter "
                    + ThisPar
                    + "."
                )
            copyIndex = processing.split("_")[1]
            cloneValue = int(processing.split("_")[2])
            targetValues = processing.split("_")[4].strip("[]")

            if "," in targetValues:
                targetList = [int(m) for m in targetValues.split(",")]
            else:
                startValue, endValue = map(int, targetValues.split(":"))
                targetList = list(range(startValue, endValue + 1))

            if copyIndex not in ThisParIx:
                Mylog.error(
                    "Copy processing error: index "
                    + copyIndex
                    + " not a dimension for parameter "
                    + ThisPar
                    + "."
                )
            if (
                cloneValue
                not in IndexTable.set_index("IndexLetter")
                .loc[copyIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Copy processing error: "
                    + cloneValue
                    + " not in the classification for aspect "
                    + copyIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )
            if not set(targetList).issubset(
                IndexTable.set_index("IndexLetter")
                .loc[copyIndex]
                .Classification.Items
            ):
                Mylog.error(
                    "Copy processing error: "
                    + str(targetList)
                    + " not entirely in the classification for aspect "
                    + copyIndex
                    + " for parameter "
                    + ThisPar
                    + "."
                )

            ix_position = ThisParIx.find(copyIndex)
            cloneIndex = (
                IndexTable.set_index("IndexLetter")
                .loc[copyIndex]
                .Classification.Items.index(cloneValue)
            )
            dimensions = Values.shape
            for indices in np.ndindex(
                dimensions[:ix_position] + dimensions[ix_position + 1 :]
            ):
                for target in targetList:
                    targetIndex = (
                        IndexTable.set_index("IndexLetter")
                        .loc[copyIndex]
                        .Classification.Items.index(target)
                    )
                    Values[
                        indices[:ix_position]
                        + (targetIndex,)
                        + indices[ix_position:]
                    ] = Values[
                        indices[:ix_position]
                        + (cloneIndex,)
                        + indices[ix_position:]
                    ]
            Mylog.info(
                "Copied  "
                + str(len(targetList))
                + " values for aspect "
                + copyIndex
                + " for parameter "
                + ThisPar
                + "."
            )

        else:
            Mylog.error(
                "Data processing error: instruction not recognized for parameter "
                + ThisPar
                + "."
            )

    if ParseUncertainty:
        return MetaData, Values, Uncertainty
    else:
        return MetaData, Values
