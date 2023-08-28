# -*- coding: utf-8 -*-
"""
Created on Thu Mar  2 17:33:00 2017

@author: spauliuk
"""

"""
File ODYM_Functions
Check https://github.com/IndEcol/ODYM for latest version.

Contains class definitions for ODYM

standard abbreviation: msf (material-system-functions)

dependencies:
    numpy >= 1.9
    scipy >= 0.14

Repository for this class, documentation, and tutorials: https://github.com/IndEcol/ODYM

"""

import os
import logging
import numpy as np
#import pandas as pd
import xlrd
import openpyxl
import pypandoc
from scipy.interpolate import make_interp_spline
from scipy.interpolate import interp1d
import ODYM_Classes as msc

####################################
#      Define functions            #
####################################

def __version__():  # return version of this file
    return str('1.0')



def function_logger(log_filename, log_pathname, file_level=logging.DEBUG, console_level=logging.WARNING):
    """
    This is the logging routine of the model. It returns alogger that can be used by other functions to write to the
    log(file).

    :param file_level: Verbosity level for the logger's output file. This can be log.WARNING (default),
        log.INFO, log.DEBUG
    :param log_filename: The filename for the logfile.
    :param log_pathname: The pathname for the logfile.
    :param console_level: Verbosity level for the logger's output file.
    out
    :param logfile_type: Type of file to write. Markdown syntax is the default.
        TODO: If other outputs types are desired, they can be converted via pandoc.
    :return: A logger that can be used by other files to write to the log(file)
    """

    log_file = os.path.join(log_pathname, log_filename)
    # logging.basicConfig(format='%(levelname)s (%(filename)s <%(funcName)s>): %(message)s',
    #                     filename=log_file,
    #                     level=logging.INFO)
    logger = logging.getLogger()
    logger.handlers = []  # required if you don't want to exit the shell
    logger.setLevel(file_level)

    # The logger for console output
    console_log = logging.StreamHandler() #StreamHandler logs to console
    console_log.setLevel(console_level)
    # console_log_format = logging.Formatter('%(message)s')
    console_log_format = logging.Formatter('%(levelname)s (%(filename)s <%(funcName)s>): %(message)s')
    console_log.setFormatter(console_log_format)
    logger.addHandler(console_log)

    # The logger for log file output
    file_log = logging.FileHandler(log_file, mode='w', encoding=None, delay=False)
    file_log.setLevel(file_level)
    file_log_format = logging.Formatter('%(message)s\n')
    file_log.setFormatter(file_log_format)
    logger.addHandler(file_log)

    return logger,  console_log, file_log



def ensure_dir(f): # Checks whether a given directory f exists, and creates it if not
    d = os.path.dirname(f)
    if not os.path.exists(d):
        os.makedirs(d)     
        
        
        
def sort_index(mylist,direction): # returns index that sorts a list, either ascending or descending
    if direction == 'ascending':
        return sorted(range(len(mylist)), key=lambda k: mylist[k])       
    elif direction == 'descending':
        return sorted(range(len(mylist)), key=lambda k: mylist[k], reverse=True)
    else:
        return None



def GroupingDict2Array(GroupingDict, ElementList):
    '''
    Tbd.
    '''
    NoOfItems = len(GroupingDict.keys())
    GroupingList = []
    for m in GroupingDict.keys():
        GroupingList.append(m)
    ElementContentArray = np.zeros((100,NoOfItems))
    PosCount = 0
    for m in GroupingList:
        for n in GroupingDict[m].keys():
            ElInd = ElementList.index(n)
            ElementContentArray[ElInd,PosCount] = GroupingDict[m][n]
        PosCount += 1
    return GroupingList, ElementContentArray



def ListStringToListNumbers(ListStr):
    """
    Extracts numbers from a string that looks like a list commant in python, and returns them as proper list
    Examples: ListStringToListNumbers('[1,2,3]') yields [1,2,3]
    """
    return [int(s) for s in ListStr[ListStr.find('['):ListStr.find(']')+1].replace('[',',').replace(']',',').split(',') if s.isdigit()]



def EvalItemSelectString(ItemSelectStr,IndexLength): 
    '''
    Extract index item selection lists from ODYM datafile information
    '''
    if ItemSelectStr == 'All' or ItemSelectStr == 'ALL' or ItemSelectStr == 'all':
        Res = 'all' # Selects all from list
    elif ItemSelectStr.find('except') > -1: # type 'All except', return full list [0,1,2,5,6,7]
        Res = np.arange(0,IndexLength)
        b = ItemSelectStr[ItemSelectStr.find('['):ItemSelectStr.find(']')+1].replace('[',',').replace(']',',')
        RemoveList = [int(s) for s in b.split(',') if s.isdigit()]   
        Res = np.delete(Res,RemoveList)      
        Res = Res.tolist() 
    elif ItemSelectStr.find(']') > -1: # type '[...]', return full list
        Res = ItemSelectStr[ItemSelectStr.find('[')::]
    elif ItemSelectStr.find(')') > -1: # type '[..:..)', return range a:b
        Res = ItemSelectStr[ItemSelectStr.find('[')+1:-1]
    else:
        Res = 'ItemSelectString could not be detected.'
    
    return Res



def MI_Tuple(value, Is): 
    """
    Define function for obtaining multiindex tuple from index value
    value: flattened index position, Is: Number of values for each index dimension
    Example: MI_Tuple(10, [3,4,2,6]) returns [0,0,1,4]
    MI_Tuple is the inverse of Tuple_MI.    
    """
    IsValuesRev = []
    CurrentValue = value
    for m in range(0,len(Is)):
        IsValuesRev.append(CurrentValue % Is[len(Is)-m-1])
        CurrentValue = CurrentValue // Is[len(Is)-m-1]
    return IsValuesRev[::-1]



def Tuple_MI(Tuple, IdxLength): 
    """
    Function to return the absolution position of a multiindex when the index tuple
    and the index hierarchy and size are given.
    Example: Tuple_MI([2,7,3],[100,10,5]) = 138
    Tuple_MI is the inverse of MI_Tuple.
    """
    # First, generate the index position offset values
    A =  IdxLength[1:] +  IdxLength[:1] # Shift 1 to left
    A[-1] = 1 # Replace lowest index by 1
    A.reverse()
    IdxPosOffset = np.cumproduct(A).tolist()
    IdxPosOffset.reverse()
    Position = np.sum([a*b for a,b in zip(Tuple,IdxPosOffset)])
    return Position


def TableWithFlowsToShares(Table,axis):
    """
    Given a 2D-table with flow values that sum up to a total, 
    either along the columns (= across rows, axis =0) or along the rows (=across the columns, axis =1).
    The function then converts the flows into shares (between 0 and 1), that each element has in the column sum (axis =0) 
    or the row sum (axis =1).
    Only makes sense if all table entries have the same sign, that is not checked by the function.
    """
    Shares = np.zeros(Table.shape)
    if axis == 0: # shares along columns
        colsum  = Table.sum(axis=0)
        Divisor = np.einsum('b,a->ab',colsum,np.ones(Table.shape[0]))
    if axis == 1: # shares along rows
        rowsum = Table.sum(axis=1)
        Divisor = np.einsum('a,b->ab',rowsum,np.ones(Table.shape[1]))
    Divided = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    Shares  = Table * Divided
    return Shares


def DetermineElementComposition_All_Oth(me):
    """
    Given an array of flows of materials (rows) broken down into chem. elements (columns), 
    where the first element is "all" and the last element is "other",
    the function determines the share of each element in the material, and fills nonexistent rows with a 1 for all and other, resp.
    """
    result        = np.zeros(me.shape)
    Shares        = TableWithFlowsToShares(me[:,1::],1)
    SharesSum     = Shares.sum(axis=1)
    result[:,0]   = 1
    result[:,1::] = Shares.copy()
    for m in range(0,me.shape[0]):
        if SharesSum[m] == 0:
            result[m,-1] = 1
    return result


def ModelIndexPositions_FromData(Positions,RowPos,ColPos):
    """
    This function is needed to read data files into ODYM. It takes the positions of a given data point 
    in the parameter file and checks where in the model index structure this data points belongs, 
    if it is needed at all.
    """
    TargetPosition = []
    for m in range(0,len(Positions)):
        if m < len(RowPos):
            try:
                TargetPosition.append(Positions[m].index(RowPos[m]))
            except:
                break
        else:
            try:
                TargetPosition.append(Positions[m].index(ColPos[m-len(RowPos)]))
            except:
                break
    return TargetPosition


def ParseModelControl(Model_Configsheet,ScriptConfig):
    """ Parse the RECC and ODYM model control parameters from the ODYM config sheet. """
    SCix = 1
    # search for script config list entry
    while Model_Configsheet.cell(SCix, 2).value != 'General Info':
        SCix += 1
            
    SCix += 2  # start on first data row
    while Model_Configsheet.cell(SCix, 4).value is not None:
        ScriptConfig[Model_Configsheet.cell(SCix, 3).value] = Model_Configsheet.cell(SCix,4).value
        SCix += 1
    
    SCix = 1
    # search for script config list entry
    while Model_Configsheet.cell(SCix, 2).value != 'Software version selection':
        SCix += 1
            
    SCix += 2 # start on first data row
    while Model_Configsheet.cell(SCix, 4).value is not None:
        ScriptConfig[Model_Configsheet.cell(SCix, 3).value] = Model_Configsheet.cell(SCix,4).value
        SCix += 1  
        
    return ScriptConfig


def ParseClassificationFile_Main(Classsheet,Mylog):
    """ Parse the ODYM classification file, format version 
    """
    ci = 2  # column index to start with
    MasterClassification = {}  # Dict of master classifications
    while Classsheet.cell(1,ci).value is not None:
        TheseItems = []
        ri = 11  # row index to start with
        ThisName = Classsheet.cell(1,ci).value
        ThisDim  = Classsheet.cell(2,ci).value
        ThisID   = Classsheet.cell(4,ci).value
        ThisUUID = Classsheet.cell(5,ci).value
        while Classsheet.cell(ri,ci).value is not None:
            TheseItems.append(Classsheet.cell(ri,ci).value) # read the classification items
            ri += 1
        MasterClassification[ThisName] = msc.Classification(Name = ThisName, Dimension = ThisDim, ID = ThisID, UUID = ThisUUID, Items = TheseItems)
        ci += 1
        
    return MasterClassification


def ParseConfigFile(Model_Configsheet,ScriptConfig,Mylog):
    """
    Standard routine to parse the ODYM model config file.
    """    
    ITix = 0
    
    # search for index table entry
    while True:
        if Model_Configsheet.cell(ITix+1, 2).value == 'Index Table':
            break
        else:
            ITix += 1
            
    IT_Aspects        = []
    IT_Description    = []
    IT_Dimension      = []
    IT_Classification = []
    IT_Selector       = []
    IT_IndexLetter    = []
    ITix += 2 # start on first data row
    while Model_Configsheet.cell(ITix+1,3).value is not None:
        IT_Aspects.append(Model_Configsheet.cell(ITix+1,3).value)
        IT_Description.append(Model_Configsheet.cell(ITix+1,4).value)
        IT_Dimension.append(Model_Configsheet.cell(ITix+1,5).value)
        IT_Classification.append(Model_Configsheet.cell(ITix+1,6).value)
        IT_Selector.append(Model_Configsheet.cell(ITix+1,7).value)
        IT_IndexLetter.append(Model_Configsheet.cell(ITix+1,8).value)        
        ITix += 1

    Mylog.info('Read parameter list from model config sheet.')
    PLix = 0
    while True: # search for parameter list entry
        if Model_Configsheet.cell(PLix+1, 2).value == 'Model Parameters':
            break
        else:
            PLix += 1
            
    PL_Names          = []
    PL_Description    = []
    PL_Version        = []
    PL_IndexStructure = []
    PL_IndexMatch     = []
    PL_IndexLayer     = []
    PL_SubFolder      = []
    PL_ProxyCode      = []
    PL_ProcMethod     = []
    PL_UpdateOverwrite = [] #2308 add choice to read new par data or use data from dat file
    
    PLix += 2 # start on first data row
    while Model_Configsheet.cell(PLix+1,3).value is not None:
        PL_Names.append(Model_Configsheet.cell(PLix+1,3).value)
        PL_Description.append(Model_Configsheet.cell(PLix+1,4).value)
        PL_Version.append(Model_Configsheet.cell(PLix+1,5).value)
        PL_IndexStructure.append(Model_Configsheet.cell(PLix+1,6).value)
        PL_IndexMatch.append(Model_Configsheet.cell(PLix+1,7).value)
        PL_IndexLayer.append(ListStringToListNumbers(Model_Configsheet.cell(PLix+1,8).value)) # strip numbers out of list string
        PL_SubFolder.append(Model_Configsheet.cell(PLix+1,12).value)
        PL_ProxyCode.append(Model_Configsheet.cell(PLix+1,13).value)
        PL_ProcMethod.append(Model_Configsheet.cell(PLix+1,14).value)
        PL_UpdateOverwrite.append(Model_Configsheet.cell(PLix+1,15).value) #2308 add choice to read new par data or use data from dat file
        PLix += 1
        
    Mylog.info('Read process list from model config sheet.')
    PrLix = 1
    
    # search for process list entry
    while Model_Configsheet.cell(PrLix, 2).value != 'Process Group List':
        PrLix += 1
            
    PrL_Number         = []
    PrL_Name           = []
    PrL_Comment        = []
    PrL_Type           = []
    PrLix += 2 # start on first data row
    
    while True:
        if Model_Configsheet.cell(PrLix,3).value is None:
            break
        PrL_Number.append(int(Model_Configsheet.cell(PrLix,3).value))
        PrL_Name.append(Model_Configsheet.cell(PrLix,4).value)
        PrL_Type.append(Model_Configsheet.cell(PrLix,5).value)
        PrL_Comment.append(Model_Configsheet.cell(PrLix,6).value)
        PrLix += 1
        
    # while Model_Configsheet.cell(PrLix,3).value is not None:
    #     print(Model_Configsheet.cell(PrLix,3).value)
    #     PrL_Number.append(int(Model_Configsheet.cell(PrLix,3).value))
    #     PrL_Name.append(Model_Configsheet.cell(PrLix,4).value)
    #     PrL_Type.append(Model_Configsheet.cell(PrLix,5).value)
    #     PrL_Comment.append(Model_Configsheet.cell(PrLix,6).value)
    #     PrLix += 1
        
    Mylog.info('Read model run control from model config sheet.')
    PrLix = 0
    
    # search for model flow control entry
    while True:
        if Model_Configsheet.cell(PrLix+1, 2).value == 'Model flow control':
            break
        else:
            PrLix += 1
    
    # start on first data row
    PrLix += 2
    while True:
        if Model_Configsheet.cell(PrLix+1, 3).value is not None:
            try:
                ScriptConfig[Model_Configsheet.cell(PrLix+1, 3).value] = Model_Configsheet.cell(PrLix+1,4).value
            except:
                None
            PrLix += 1
        else:
            break  
    
    Mylog.info('Read model output control from model config sheet.')
    PrLix = 0
    
    # search for model flow control entry
    while True:
        if Model_Configsheet.cell(PrLix+1, 2).value == 'Model output control':
            break
        else:
            PrLix += 1
    
    # start on first data row
    PrLix += 2
    while True:
        if Model_Configsheet.cell(PrLix+1, 3).value is not None:
            try:
                ScriptConfig[Model_Configsheet.cell(PrLix+1, 3).value] = Model_Configsheet.cell(PrLix+1,4).value
            except:
                None
            PrLix += 1
        else:
            break  
    
    return IT_Aspects,IT_Description,IT_Dimension,IT_Classification,IT_Selector,IT_IndexLetter,PL_Names,PL_Description,PL_Version,PL_IndexStructure,PL_IndexMatch,PL_IndexLayer,PL_SubFolder,PL_ProxyCode,PL_ProcMethod,PL_UpdateOverwrite,PrL_Number,PrL_Name,PrL_Comment,PrL_Type,ScriptConfig


def ReadParameter(ParPath, ThisPar, ThisParIx, IndexMatch, ThisParLayerSel, MasterClassification,
                  IndexTable, IndexTable_ClassificationNames, ScriptConfig, Mylog):
    """
    This function reads a model parameter from the corresponding parameter file
    """
    Parfile   = xlrd.open_workbook(ParPath + '.xlsx')
    ParHeader = Parfile.sheet_by_name('Cover')
    
    IM = eval(IndexMatch) # List that matches model aspects to parameter indices
    
    ri = 1 # row index
    MetaData = {}
    while True: # read cover sheet info
        ThisItem = ParHeader.cell_value(ri,0)
        if ThisItem != 'Dataset_RecordType':
            MetaData[ThisItem] = ParHeader.cell_value(ri,1)
            ri += 1
        else:
            break # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType
    
    # Check whether parameter file uses same classification:
    if 'ODYM_Classifications_Master_' + \
            ScriptConfig['Version of master classification'] != MetaData['Dataset_Classification_version_number']:
        Mylog.critical('CLASSIFICATION FILE FATAL ERROR: Classification file of parameter ' + ThisPar +
                       ' is not identical to the classification master file used for the current model run.')
        
    if ParHeader.cell_value(ri,1) == 'List':
        IList = []
        IListMeaning = []
        ci = 1 # column index
        while True:
            if ParHeader.cell_value(ri +1,ci) != '':
                IList.append(ParHeader.cell_value(ri +1,ci))
                IListMeaning.append(ParHeader.cell_value(ri +2,ci))
                ci += 1
            else:
                break
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]
            
        ValueList = []
        VIComment = []
        ci = 1 # column index
        while True:
            if ParHeader.cell_value(ri +4,ci) != '':
                ValueList.append(ParHeader.cell_value(ri +3,ci))
                VIComment.append(ParHeader.cell_value(ri +4,ci))
                ci += 1
            else:
                break
        
        # Check whether all indices are present in the index table of the model  
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Index list of data file for parameter ' + ThisPar +
                        ' contains indices that are not part of the current model run.')
    
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != IList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' + IList[m])
                break  # Stop parsing parameter, will cause model to halt
            
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])

        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))
        ValIns = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name('Values_Master')
        ColOffset = len(IList)
        RowOffset = 1 # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx        = 0
        while True:
            try:
                CV = ValuesSheet.cell_value(cx + RowOffset, ColOffset)
            except:
                break
            TargetPosition = []
            for mx in range(0,len(IList)): # mx iterates over the aspects of the parameter 
                CurrentItem = ValuesSheet.cell_value(cx + RowOffset, IM[mx])
                try:
                    TargetPosition.append(IndexTable.set_index('IndexLetter').loc[ThisParIx[mx]].Classification.Items.index(CurrentItem))
                except:
                    break # Current parameter value is not needed for model, outside scope for a certain aspect. 
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
            cx += 1
            
        Mylog.info('A total of ' + str(cx+1) + ' values was read from file for parameter ' + ThisPar + '.')
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar + ' were assigned.')
         
        
        
    ### Table version ###
    if ParHeader.cell_value(ri,1) == 'Table': # have 3 while loops, one for row indices, one for column indices, one for value layers
       
        RIList        = []
        RISize        = []
        RIListMeaning = []
        ci = 1 # column index
        while True:
            if ParHeader.cell_value(ri +1,ci) != '':
                RIList.append(ParHeader.cell_value(ri +1,ci))
                RISize.append(int(ParHeader.cell_value(ri +2,1)))
                RIListMeaning.append(ParHeader.cell_value(ri +3,ci))
                ci += 1
            else:
                break
        RISize = RISize[0]      
            
        CIList        = []
        CISize        = []
        CIListMeaning = []
        ci = 1 # column index
        while True:
            if ParHeader.cell_value(ri +4,ci) != '':
                CIList.append(ParHeader.cell_value(ri +4,ci))
                CISize.append(int(ParHeader.cell_value(ri +5,1)))
                CIListMeaning.append(ParHeader.cell_value(ri +6,ci))
                ci += 1
            else:
                break
        CISize = CISize[0]
        
        # Re-Order indices to fit model aspect order:
        ComIList        = RIList        + CIList    
        ComIList        = [ComIList[i] for i in IM]                
            
        ValueList = []
        VIComment = []
        ci = 1 # column index
        while True:
            if ParHeader.cell_value(ri +7,ci) != '':
                ValueList.append(ParHeader.cell_value(ri +7,ci))
                VIComment.append(ParHeader.cell_value(ri +8,ci))
                ci += 1
            else:
                break
        
        # Check whether all indices are present in the index table of the model  
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Row index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Column index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
            
        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0,len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])    
        CIIndexLetter = []
        for m in range(0,len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m+len(RIList))])    
        
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' +
                            ComIList[m])
                break  # Stop parsing parameter, will cause model to halt
                
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])
        
        # Read parameter values into array:
        Values = np.zeros((IndexSizesM))
        ValIns = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name(ValueList[ThisParLayerSel[0]])
        ColOffset = len(RIList)
        RowOffset = len(CIList)
        RowNos    = RISize   
        ColNos    = CISize
        
        TargetPos_R = []
        for m in range(0,RowNos):
            TP_RD = []
            for mc in range(0,len(RIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(m + RowOffset, mc))
                except:
                    CurrentItem = ValuesSheet.cell_value(m + RowOffset, mc)
                try:
                    IX   = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[RIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_RD.append((IX,TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)           

        TargetPos_C = []                
        for n in range(0,ColNos):
            TP_CD = []
            for mc in range(0,len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(mc, n + ColOffset))
                except:
                    CurrentItem = ValuesSheet.cell_value(mc, n + ColOffset)
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[CIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_CD.append((IX,TPIX))
                except:
                    TP_CD.append(None)
                    break  
            TargetPos_C.append(TP_CD)
        
        for m in range(0,RowNos):
            for n in range(0,ColNos):
                TargetPosition = [0 for i in range(0,len(ComIList))]
                try:
                    for i in range(0,len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1] 
                    for i in range(0,len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1] 
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(ComIList):
                    Values[tuple(TargetPosition)] = ValuesSheet.cell_value(m + RowOffset, n + ColOffset)
                    ValIns[tuple(TargetPosition)] = 1
                    
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar +
                   ' were assigned.')
       
    return MetaData, Values



def ReadParameterV2(ParPath, ThisPar, ThisParIx, IndexMatch, ThisParLayerSel, MasterClassification,
                    IndexTable, IndexTable_ClassificationNames, ScriptConfig, Mylog, ParseUncertainty):
    """
    This function reads a model parameter from the corresponding parameter file
    """
    Parfile   = xlrd.open_workbook(ParPath + '.xlsx')
    ParHeader = Parfile.sheet_by_name('Cover')
    
    IM = eval(IndexMatch) # List that matches model aspects to parameter indices
    
    ri = 1 # row index
    MetaData = {}
    while True: # read cover sheet info
        ThisItem = ParHeader.cell_value(ri,0)
        if (ThisItem != '[Empty on purpose]' and ThisItem != 'Dataset_RecordType'):
            MetaData[ThisItem] = ParHeader.cell_value(ri,1)
            if ThisItem == 'Dataset_Unit':
                if ParHeader.cell_value(ri,1) == 'GLOBAL':
                    MetaData['Unit_Global']         = ParHeader.cell_value(ri,2)
                    MetaData['Unit_Global_Comment'] = ParHeader.cell_value(ri,3) 
            if ThisItem == 'Dataset_Uncertainty':
                # if LIST is specified, nothing happens here.
                if ParHeader.cell_value(ri,1) == 'GLOBAL':
                    MetaData['Dataset_Uncertainty_Global'] = ParHeader.cell_value(ri,2)
                if ParHeader.cell_value(ri,1) == 'TABLE':
                    MetaData['Dataset_Uncertainty_Sheet']  = ParHeader.cell_value(ri,2)                    
            if ThisItem == 'Dataset_Comment':
                if ParHeader.cell_value(ri,1) == 'GLOBAL':
                    MetaData['Dataset_Comment_Global']     = ParHeader.cell_value(ri,2)                    
            ri += 1
        else:
            break # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType
    
    # Check whether parameter file uses same classification:
    if  ScriptConfig['Version of master classification'] != MetaData['Dataset_Classification_version_number']:
        Mylog.critical('CLASSIFICATION FILE FATAL ERROR: Classification file of parameter ' + ThisPar +
                       ' is not identical to the classification master file used for the current model run.')
        
    # Continue parsing until line 'Dataset_RecordType' is found:
    while True:
        ThisItem = ParHeader.cell_value(ri,0)
        if ThisItem == 'Dataset_RecordType':  
            break
        else:
            ri += 1
        
    ### List version ###        
    if ParHeader.cell_value(ri,1) == 'LIST':
        IList = []
        IListMeaning = []
        RI_Start = ri + 2
        while True:
            if ParHeader.cell_value(RI_Start,0) != '':
                IList.append(ParHeader.cell_value(RI_Start,0))
                IListMeaning.append(ParHeader.cell_value(RI_Start,1))
                RI_Start += 1
            else:
                break
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]
            
        ValueList = []
        VIComment = []
        RI_Start = ri + 2
        while True:
            if ParHeader.cell_value(RI_Start,2) != '':
                ValueList.append(ParHeader.cell_value(RI_Start,2))
                VIComment.append(ParHeader.cell_value(RI_Start,3))
                RI_Start += 1
            else:
                break
        
        # Check whether all indices are present in the index table of the model  
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Index list of data file for parameter ' + ThisPar +
                        ' contains indices that are not part of the current model run.')
    
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != IList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' + IList[m])
                break  # Stop parsing parameter, will cause model to halt
            
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])

        # Read parameter values into array, uncertainty into list:
        Values      = np.zeros((IndexSizesM)) # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM) # parameter value uncertainties  
        ValIns      = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded
        ValuesSheet = Parfile.sheet_by_name('Values_Master')
        ColOffset = len(IList)
        RowOffset = 1 # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx        = 0
        while True:
            try:
                CV = ValuesSheet.cell_value(cx + RowOffset, ColOffset)
            except:
                break
            TargetPosition = []
            for mx in range(0,len(IList)): # mx iterates over the aspects of the parameter 
                CurrentItem = ValuesSheet.cell_value(cx + RowOffset, IM[mx])
                try:
                    TargetPosition.append(IndexTable.set_index('IndexLetter').loc[ThisParIx[mx]].Classification.Items.index(CurrentItem))
                except:
                    break # Current parameter value is not needed for model, outside scope for a certain aspect. 
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
                Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = ValuesSheet.cell_value(cx + RowOffset, ColOffset + 3)
            cx += 1
            
        Mylog.info('A total of ' + str(cx) + ' values was read from file for parameter ' + ThisPar + '.')
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar + ' were assigned.')
         
        
        
    ### Table version ###
    if ParHeader.cell_value(ri,1) == 'TABLE': # have 3 while loops, one for row indices, one for column indices, one for value layers
        ColNos =  int(ParHeader.cell_value(ri,5)) # Number of columns in dataset
        RowNos =  int(ParHeader.cell_value(ri,3)) # Number of rows in dataset
        
        RI = ri + 2 # row where indices start
        RIList        = []
        RIListMeaning = []
        while True:
            if ParHeader.cell_value(RI,0) != '':
                RIList.append(ParHeader.cell_value(RI,0))
                RIListMeaning.append(ParHeader.cell_value(RI,1))
                RI += 1
            else:
                break

        RI = ri + 2 # row where indices start    
        CIList        = []
        CIListMeaning = []
        while True:
            if ParHeader.cell_value(RI,2) != '':
                CIList.append(ParHeader.cell_value(RI,2))
                CIListMeaning.append(ParHeader.cell_value(RI,3))
                RI += 1
            else:
                break
        
        # Re-Order indices to fit model aspect order:
        ComIList        = RIList        + CIList    # List of all indices, both rows and columns
        ComIList        = [ComIList[i] for i in IM]                
            
        RI = ri + 2 # row where indices start  
        ValueList = []
        VIComment = []
        while True:
            if ParHeader.cell_value(RI,4) != '':
                ValueList.append(ParHeader.cell_value(RI,4))
                VIComment.append(ParHeader.cell_value(RI,5))
                RI += 1
            else:
                break
        
        # Check whether all indices are present in the index table of the model  
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Row index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Column index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
            
        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0,len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])    
        CIIndexLetter = []
        for m in range(0,len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m+len(RIList))])    
        
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' +
                            ComIList[m])
                break  # Stop parsing parameter, will cause model to halt
                
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])
        
        # Read parameter values into array:
        Values      = np.zeros((IndexSizesM)) # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM) # parameter value uncertainties  
        ValIns      = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded, contains 0 or 1.
        ValuesSheet = Parfile.sheet_by_name(ValueList[ThisParLayerSel[0]])
        if ParseUncertainty == True:
            if 'Dataset_Uncertainty_Sheet' in MetaData:
                UncertSheet = Parfile.sheet_by_name(MetaData['Dataset_Uncertainty_Sheet'])
        ColOffset   = len(RIList)
        RowOffset   = len(CIList)
        cx          = 0
        
        TargetPos_R = [] # Determine all row target positions in data array
        for m in range(0,RowNos):
            TP_RD = []
            for mc in range(0,len(RIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(m + RowOffset, mc)) # in case items come as int, e.g., years
                except:
                    CurrentItem = ValuesSheet.cell_value(m + RowOffset, mc)
                try:
                    IX   = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[RIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_RD.append((IX,TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)         
                

        TargetPos_C = [] # Determine all col target positions in data array  
        for n in range(0,ColNos):
            TP_CD = []
            for mc in range(0,len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell_value(mc, n + ColOffset))
                except:
                    CurrentItem = ValuesSheet.cell_value(mc, n + ColOffset)
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[CIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_CD.append((IX,TPIX))
                except:
                    TP_CD.append(None)
                    break  
            TargetPos_C.append(TP_CD)
        
        for m in range(0,RowNos): # Read values from excel template
            for n in range(0,ColNos):
                TargetPosition = [0 for i in range(0,len(ComIList))]
                try:
                    for i in range(0,len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1] 
                    for i in range(0,len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1] 
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(ComIList): # Read value if TargetPosition Tuple has same length as indexList
                    Values[tuple(TargetPosition)] = ValuesSheet.cell_value(m + RowOffset, n + ColOffset)
                    ValIns[tuple(TargetPosition)] = 1
                    # Add uncertainty
                    if ParseUncertainty == True:
                        if 'Dataset_Uncertainty_Global' in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = MetaData['Dataset_Uncertainty_Global']
                        if 'Dataset_Uncertainty_Sheet' in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = UncertSheet.cell_value(m + RowOffset, n + ColOffset)
                cx += 1

        Mylog.info('A total of ' + str(cx) + ' values was read from file for parameter ' + ThisPar + '.')                    
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar +
                   ' were assigned.')
    if ParseUncertainty == True:
        return MetaData, Values, Uncertainty
    else:
        return MetaData, Values
    
def ReadParameterXLSX(ParPath, ThisPar, ThisParIx, IndexMatch, ThisParLayerSel, ThisParProcMethod, MasterClassification,
                    IndexTable, IndexTable_ClassificationNames, ScriptConfig, Mylog, ParseUncertainty):
    """
    This function reads a model parameter from the corresponding parameter file and used openpyxl
    """
    Parfile   = openpyxl.load_workbook(ParPath + '.xlsx', data_only=True)
    ParHeader = Parfile['Cover']
    
    IM = eval(IndexMatch) # List that matches model aspects to parameter indices
    
    ri = 2 # row index
    MetaData = {}
    while True: # read cover sheet info
        ThisItem = ParHeader.cell(ri,1).value
        if (ThisItem != '[Empty on purpose]' and ThisItem != 'Dataset_RecordType'):
            MetaData[ThisItem] = ParHeader.cell(ri,2).value
            if ThisItem == 'Dataset_Unit':
                if ParHeader.cell(ri,2).value == 'GLOBAL':
                    MetaData['Unit_Global']         = ParHeader.cell(ri,3).value
                    MetaData['Unit_Global_Comment'] = ParHeader.cell(ri,4).value 
            if ThisItem == 'Dataset_Uncertainty':
                # if LIST is specified, nothing happens here.
                if ParHeader.cell(ri,2).value == 'GLOBAL':
                    MetaData['Dataset_Uncertainty_Global'] = ParHeader.cell(ri,3).value
                if ParHeader.cell(ri,2).value == 'TABLE':
                    MetaData['Dataset_Uncertainty_Sheet']  = ParHeader.cell(ri,3).value                    
            if ThisItem == 'Dataset_Comment':
                if ParHeader.cell(ri,2).value == 'GLOBAL':
                    MetaData['Dataset_Comment_Global']     = ParHeader.cell(ri,3).value                    
            ri += 1
        else:
            break # terminate while loop when all meta information is read.
            # Now we are in the row of Dataset_RecordType
    
    # Check whether parameter file uses same classification:
    if  ScriptConfig['Version of master classification'] != MetaData['Dataset_Classification_version_number']:
        Mylog.critical('CLASSIFICATION FILE FATAL ERROR: Classification file of parameter ' + ThisPar +
                       ' is not identical to the classification master file used for the current model run.')
        
    # Continue parsing until line 'Dataset_RecordType' is found:
    while True:
        ThisItem = ParHeader.cell(ri,1).value
        if ThisItem == 'Dataset_RecordType':  
            Mylog.info(ParHeader.cell(ri,2).value)
            break
        else:
            ri += 1
        
    ### List version ###        
    if ParHeader.cell(ri,2).value == 'LIST': # ri = 21
        IList = []
        IListMeaning = []
        RI_Start = ri + 2
        while ParHeader.cell(RI_Start,1).value is not None:
            IList.append(ParHeader.cell(RI_Start,1).value)
            IListMeaning.append(ParHeader.cell(RI_Start,2).value)
            RI_Start += 1
        # Re-Order indices to fit model aspect order:
        IList = [IList[i] for i in IM]
        IListMeaning = [IListMeaning[i] for i in IM]
            
        ValueList = []
        VIComment = []
        RI_Start = ri + 2
        while ParHeader.cell(RI_Start,3).value is not None:
            ValueList.append(ParHeader.cell(RI_Start,3).value)
            VIComment.append(ParHeader.cell(RI_Start,4).value)
            RI_Start += 1
        
        # Check whether all indices are present in the index table of the model  
        if set(IList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Index list of data file for parameter ' + ThisPar +
                        ' contains indices that are not part of the current model run.')
    
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            # Check whether index is present in parameter file:
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != IList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' + IList[m])
                break  # Stop parsing parameter, will cause model to halt
            
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])
        # Read parameter values into array, uncertainty into list:
        Values      = np.zeros((IndexSizesM)) # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM) # parameter value uncertainties  
        ValIns      = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded
        ValuesSheet = Parfile['Values_Master']
        ColOffset = len(IList)
        RowOffset = 1 # fixed for this format, different quantification layers (value, error, etc.) will be read later
        cx        = 0
        while True:
            if ValuesSheet.cell(cx + RowOffset+1, ColOffset+1).value is not None:
                CV = ValuesSheet.cell(cx + RowOffset+1, ColOffset+1).value
            else:
                break
            TargetPosition = []
            for mx in range(0,len(IList)): # mx iterates over the aspects of the parameter 
                CurrentItem = ValuesSheet.cell(cx + RowOffset+1, IM[mx]+1).value

                try:
                    TargetPosition.append(IndexTable.set_index('IndexLetter').loc[ThisParIx[mx]].Classification.Items.index(CurrentItem))
                except:
                    break # Current parameter value is not needed for model, outside scope for a certain aspect. 
            if len(TargetPosition) == len(ThisParIx):
                Values[tuple(TargetPosition)] = CV
                ValIns[tuple(TargetPosition)] = 1
                Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = ValuesSheet.cell(cx + RowOffset+1, ColOffset + 4).value
            cx += 1
            
        Mylog.info('A total of ' + str(cx) + ' values was read from file for parameter ' + ThisPar + '.')
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar + ' were assigned.')
         
        
        
    ### Table version ###
    if ParHeader.cell(ri,2).value == 'TABLE': # have 3 while loops, one for row indices, one for column indices, one for value layers
        ColNos =  int(ParHeader.cell(ri,6).value) # Number of columns in dataset
        RowNos =  int(ParHeader.cell(ri,4).value) # Number of rows in dataset
        
        RI = ri + 2 # row where indices start
        RIList        = []
        RIListMeaning = []
        while True:
            if ParHeader.cell(RI,1).value is not None:
                RIList.append(ParHeader.cell(RI,1).value)
                RIListMeaning.append(ParHeader.cell(RI,2).value)
                RI += 1
            else:
                break

        RI = ri + 2 # row where indices start    
        CIList        = []
        CIListMeaning = []
        while True:
            if ParHeader.cell(RI,3).value is not None:
                CIList.append(ParHeader.cell(RI,3).value)
                CIListMeaning.append(ParHeader.cell(RI,4).value)
                RI += 1
            else:
                break
        
        # Re-Order indices to fit model aspect order:
        ComIList        = RIList        + CIList    # List of all indices, both rows and columns
        ComIList        = [ComIList[i] for i in IM]                
            
        RI = ri + 2 # row where indices start  
        ValueList = []
        VIComment = []
        while True:
            if ParHeader.cell(RI,5).value is not None:
                ValueList.append(ParHeader.cell(RI,5).value)
                VIComment.append(ParHeader.cell(RI,6).value)
                RI += 1
            else:
                break
        
        # Check whether all indices are present in the index table of the model  
        if set(RIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Row index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
        if set(CIList).issubset(set(IndexTable_ClassificationNames)) is False:
            Mylog.error('CLASSIFICATION ERROR: Column index list of data file for parameter ' + ThisPar + ' contains indices that are not part of the current model run.')
            
        # Determine index letters for RIList and CIList
        RIIndexLetter = []
        for m in range(0,len(RIList)):
            RIIndexLetter.append(ThisParIx[IM.index(m)])    
        CIIndexLetter = []
        for m in range(0,len(CIList)):
            CIIndexLetter.append(ThisParIx[IM.index(m+len(RIList))])    
        
        # Check how well items match between model and data, select items to import
        IndexSizesM  = [] # List of dimension size for model
        for m in range(0,len(ThisParIx)):
            ThisDim = ThisParIx[m]
            ThisDimClassificationName  = IndexTable.set_index('IndexLetter').loc[ThisDim].Classification.Name
            if ThisDimClassificationName != ComIList[m]:
                Mylog.error('CLASSIFICATION ERROR: Classification ' + ThisDimClassificationName + ' for aspect ' +
                            ThisDim + ' of parameter ' + ThisPar +
                            ' must be identical to the specified classification of the corresponding parameter dimension, which is ' +
                            ComIList[m])
                break  # Stop parsing parameter, will cause model to halt
                
            IndexSizesM.append(IndexTable.set_index('IndexLetter').loc[ThisDim]['IndexSize'])
        
        # Read parameter values into array:
        Values      = np.zeros((IndexSizesM)) # Array for parameter values
        Uncertainty = [None] * np.product(IndexSizesM) # parameter value uncertainties  
        ValIns      = np.zeros((IndexSizesM)) # Array to check how many values are actually loaded, contains 0 or 1.
        ValuesSheet = Parfile[ValueList[ThisParLayerSel[0]]]
        if ParseUncertainty == True:
            if 'Dataset_Uncertainty_Sheet' in MetaData:
                UncertSheet = Parfile[MetaData['Dataset_Uncertainty_Sheet']]
        ColOffset   = len(RIList)
        RowOffset   = len(CIList)
        cx          = 0
        
        TargetPos_R = [] # Determine all row target positions in data array
        for m in range(0,RowNos):
            TP_RD = []
            for mc in range(0,len(RIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell(m + RowOffset+1, mc+1).value) # in case items come as int, e.g., years
                except:
                    CurrentItem = ValuesSheet.cell(m + RowOffset+1, mc+1).value
                try:
                    IX   = ThisParIx.find(RIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[RIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_RD.append((IX,TPIX))
                except:
                    TP_RD.append(None)
                    break
            TargetPos_R.append(TP_RD)         
                

        TargetPos_C = [] # Determine all col target positions in data array  
        for n in range(0,ColNos):
            TP_CD = []
            for mc in range(0,len(CIList)):
                try:
                    CurrentItem = int(ValuesSheet.cell(mc+1, n + ColOffset+1).value)
                except:
                    CurrentItem = ValuesSheet.cell(mc+1, n + ColOffset+1).value
                try:
                    IX = ThisParIx.find(CIIndexLetter[mc])
                    TPIX = IndexTable.set_index('IndexLetter').loc[CIIndexLetter[mc]].Classification.Items.index(CurrentItem)
                    TP_CD.append((IX,TPIX))
                except:
                    TP_CD.append(None)
                    break  
            TargetPos_C.append(TP_CD)
        
        for m in range(0,RowNos): # Read values from excel template
            for n in range(0,ColNos):
                TargetPosition = [0 for i in range(0,len(ComIList))]
                try:
                    for i in range(0,len(RIList)):
                        TargetPosition[TargetPos_R[m][i][0]] = TargetPos_R[m][i][1] 
                    for i in range(0,len(CIList)):
                        TargetPosition[TargetPos_C[n][i][0]] = TargetPos_C[n][i][1] 
                except:
                    TargetPosition = [0]
                if len(TargetPosition) == len(ComIList): # Read value if TargetPosition Tuple has same length as indexList
                    Values[tuple(TargetPosition)] = ValuesSheet.cell(m + RowOffset+1, n + ColOffset+1).value
                    ValIns[tuple(TargetPosition)] = 1
                    # Add uncertainty
                    if ParseUncertainty == True:
                        if 'Dataset_Uncertainty_Global' in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = MetaData['Dataset_Uncertainty_Global']
                        if 'Dataset_Uncertainty_Sheet' in MetaData:
                            Uncertainty[Tuple_MI(TargetPosition, IndexSizesM)] = UncertSheet.cell_value(m + RowOffset +1, n + ColOffset +1)
                cx += 1

        Mylog.info('A total of ' + str(cx) + ' values was read from file for parameter ' + ThisPar + '.')                    
        Mylog.info(str(ValIns.sum()) + ' of ' + str(np.prod(IndexSizesM)) + ' values for parameter ' + ThisPar +
                   ' were assigned.')
        
        Processing_methods = eval(ThisParProcMethod)
        for processing in Processing_methods:
        
            if processing == 'none':
                continue
                
            elif processing.startswith('replicate'):
                if len(ThisParProcMethod.split('_')) != 5:
                    Mylog.error('Replicate processing error: instruction not recognized for parameter '+ ThisPar + '.')
                
                replicateIndex = processing.split('_')[1]
                targetValue    = processing.split('_')[2]
                copyValue      = processing.split('_')[4]
                
                if replicateIndex not in ThisParIx:
                    Mylog.error('Replicate processing error: index ' + replicateIndex + ' not a dimension for parameter '+ ThisPar + '.')
                if copyValue not in IndexTable.set_index('IndexLetter').loc[replicateIndex].Classification.Items:
                    Mylog.error('Replicate processing error: ' + copyValue   + ' not in the classification for aspect ' + replicateIndex + ' for parameter '+ ThisPar + '.')
                if targetValue not in IndexTable.set_index('IndexLetter').loc[replicateIndex].Classification.Items:
                    Mylog.error('Replicate processing error: ' + targetValue + ' not in the classification for aspect ' + replicateIndex + ' for parameter '+ ThisPar + '.')
                    
                ix_position = ThisParIx.find(replicateIndex)
                C_ix = IndexTable.set_index('IndexLetter').loc[replicateIndex].Classification.Items.index(copyValue)
                T_ix = IndexTable.set_index('IndexLetter').loc[replicateIndex].Classification.Items.index(targetValue)
                dimensions = Values.shape
                for indices in np.ndindex(dimensions[:ix_position] + dimensions[ix_position + 1:]):
                    Values[indices[:ix_position] + (T_ix,) + indices[ix_position:]] = Values[indices[:ix_position] + (C_ix,) + indices[ix_position:]]
                Mylog.info('Replicated ' + copyValue + ' values in ' + targetValue + ' for aspect ' + replicateIndex + ' for parameter '+ ThisPar + '.')
    
    
            elif processing.startswith('interpolate'):
                if len(processing.split('_')) != 5:
                    Mylog.error('Interpolate processing error: instruction not recognized for parameter '+ ThisPar + '.')
                interpIndex = processing.split('_')[1]
                startValue  = int(processing.split('_')[2])
                endValue    = int(processing.split('_')[3])
                method      = processing.split('_')[4]
                  
                if interpIndex not in ThisParIx:
                    Mylog.error('Interpolation processing error: index ' + interpIndex + ' not a dimension for parameter '+ ThisPar + '.')
                if startValue not in IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items:
                    Mylog.error('Interpolation processing error: ' + str(startValue) + ' not in the classification for aspect ' + interpIndex + ' for parameter '+ ThisPar + '.')
                if endValue not in IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items:
                    Mylog.error('Interpolation processing error: ' + str(endValue)   + ' not in the classification for aspect ' + interpIndex + ' for parameter '+ ThisPar + '.')
                    
                startIndex = IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items.index(startValue)
                endIndex   = IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items.index(endValue)
                ix_position = ThisParIx.find(interpIndex)
                ValIns_b = np.array(ValIns, dtype=bool)
                dimensions = Values.shape
                
                for indices in np.ndindex(dimensions[:ix_position] + dimensions[ix_position + 1:]):
                    if (ValIns_b[indices[:ix_position] + (startIndex,) + indices[ix_position:]] and ValIns_b[indices[:ix_position] + (endIndex,) + indices[ix_position:]]):
                        x = [IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items[m] for m in range(startIndex, endIndex+1) if ValIns_b[indices[:ix_position] + (m,) + indices[ix_position:]] ]
                        y = [Values[indices[:ix_position] + (m,) + indices[ix_position:]] for m in range(startIndex, endIndex+1) if ValIns_b[indices[:ix_position] + (m,) + indices[ix_position:]] ]
                        if method == 'spline':
                            clamped_spline = make_interp_spline(x, y, bc_type=([(2, 0)], [(1, 0)])) #spline function, free (2nd derivative=0) for starting boundary condition and clamped (1st derivative=0) for end boundary condition
                            for m in range(startIndex, endIndex+1):
                                Values[indices[:ix_position] + (m,) + indices[ix_position:]] = clamped_spline(IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items[m])
                        elif method == 'linear':
                            f = interp1d(x, y, kind='linear')
                            for m in range(startIndex, endIndex+1):
                                Values[indices[:ix_position] + (m,) + indices[ix_position:]] = f(IndexTable.set_index('IndexLetter').loc[interpIndex].Classification.Items[m])
                        else:
                            Mylog.error('Interpolation error: method ' + method   + ' not recognized for parameter '+ ThisPar + '.')
                            break
                    
                Mylog.info('Intrpolated ' + str(interpIndex) + ' aspect from ' + str(startValue) + ' to ' + str(endValue) + ' for parameter ' + ThisPar + '.')
                count_neg = (Values<0).sum()   
                if count_neg >0:
                    Values[Values<0]=0
                    Mylog.info(str(count_neg) + ' negative values from spline interpolation set to 0.')
                
                
            elif processing.startswith('copy'): 
                if len(processing.split('_')) != 5:
                    Mylog.error('Copy processing error: instruction not recognized for parameter '+ ThisPar + '.')
                copyIndex    = processing.split('_')[1]
                cloneValue    = int(processing.split('_')[2])
                targetValues = processing.split('_')[4].strip('[]')
                
                if ',' in targetValues:
                    targetList = [int(m) for m in targetValues.split(',')]
                else:
                    startValue, endValue = map(int, targetValues.split(':'))
                    targetList = list(range(startValue, endValue + 1))
                
                if copyIndex not in ThisParIx:
                    Mylog.error('Copy processing error: index ' + copyIndex + ' not a dimension for parameter '+ ThisPar + '.')
                if cloneValue not in IndexTable.set_index('IndexLetter').loc[copyIndex].Classification.Items:
                    Mylog.error('Copy processing error: ' + cloneValue + ' not in the classification for aspect ' + copyIndex + ' for parameter '+ ThisPar + '.')
                if not set(targetList).issubset(IndexTable.set_index('IndexLetter').loc[copyIndex].Classification.Items):
                    Mylog.error('Copy processing error: ' + str(targetList)   + ' not entirely in the classification for aspect ' + copyIndex + ' for parameter '+ ThisPar + '.')
                
                ix_position = ThisParIx.find(copyIndex)
                cloneIndex = IndexTable.set_index('IndexLetter').loc[copyIndex].Classification.Items.index(cloneValue)
                dimensions = Values.shape
                for indices in np.ndindex(dimensions[:ix_position] + dimensions[ix_position + 1:]):
                    for target in targetList:
                        targetIndex = IndexTable.set_index('IndexLetter').loc[copyIndex].Classification.Items.index(target)
                        Values[indices[:ix_position] + (targetIndex,) + indices[ix_position:]] = Values[indices[:ix_position] + (cloneIndex,) + indices[ix_position:]]
                Mylog.info('Copied  ' + str(len(targetList)) + ' values for aspect ' + copyIndex + ' for parameter ' + ThisPar + '.')
                    
                
            else:
                Mylog.error('Data processing error: instruction not recognized for parameter '+ ThisPar + '.')
            
                
    if ParseUncertainty == True:
        return MetaData, Values, Uncertainty
    else:
        return MetaData, Values
    

def ExcelSheetFill(Workbook, Sheetname, values, topcornerlabel=None,
                   rowlabels=None, collabels=None, Style=None,
                   rowselect=None, colselect=None):
    Sheet = Workbook.add_sheet(Sheetname)
    if topcornerlabel is not None:
        if Style is not None:
            Sheet.write(0,0,label = topcornerlabel, style = Style)  # write top corner label
        else:
            Sheet.write(0,0,label = topcornerlabel)  # write top corner label
    if rowselect is None: # assign row select if not present (includes all rows in that case)
        rowselect = np.ones((values.shape[0]))
    if colselect is None: # assign col select if not present (includes all columns in that case)
        colselect = np.ones((values.shape[1]))        
    if rowlabels is not None: # write row labels
         rowindexcount = 0
         for m in range(0,len(rowlabels)):
             if rowselect[m] == 1: # True if True or 1
                 if Style is None:
                     Sheet.write(rowindexcount +1, 0, label = rowlabels[m])
                 else:
                     Sheet.write(rowindexcount +1, 0, label = rowlabels[m], style = Style)
                 rowindexcount += 1
    if collabels is not None: # write column labels
         colindexcount = 0
         for m in range(0,len(collabels)):
             if colselect[m] == 1: # True if True or 1
                 if Style is None:
                     Sheet.write(0, colindexcount +1, label = collabels[m])
                 else:
                     Sheet.write(0, colindexcount +1, label = collabels[m], style = Style)
                 colindexcount += 1   
    # write values:
    rowindexcount = 0
    for m in range(0,values.shape[0]): # for all rows
        if rowselect[m] == 1:
            colindexcount = 0
            for n in range(0,values.shape[1]): # for all columns
                if colselect[n] == 1:
                    Sheet.write(rowindexcount +1, colindexcount + 1, label=values[m, n])
                    colindexcount += 1
            rowindexcount += 1
                       
def ExcelExportAdd_tAB(Sheet,Data,rowoffset,coloffset,IName,UName,RName,FName,REName,ALabels,BLabels):
    """
    This function exports a 3D array with aspects time, A, and B to a given excel sheet.
    Same as xlsxExportAdd_tAB but this function is for xls files with xlrd.
    The t dimension is exported in one row, the A and B dimensions as several rows.
    Each row starts with IName (indicator), UName (unit), RName (region), 
    FName (figure where data are used), REName (Resource efficiency scenario), 
    and then come the values for the dimensions A and B and from coloffset onwards, the time dimension.
    Function is meant to be used multiple times, so a rowoffset is given, incremented, and returned for the next run.
    """
    for m in range(0,len(ALabels)):
        for n in range(0,len(BLabels)):
            Sheet.write(rowoffset, 0, label = IName)
            Sheet.write(rowoffset, 1, label = UName)
            Sheet.write(rowoffset, 2, label = RName)
            Sheet.write(rowoffset, 3, label = FName)
            Sheet.write(rowoffset, 4, label = REName)
            Sheet.write(rowoffset, 5, label = ALabels[m])
            Sheet.write(rowoffset, 6, label = BLabels[n])
            for t in range(0,Data.shape[0]):
                Sheet.write(rowoffset, coloffset + t, label = Data[t,m,n])
            rowoffset += 1
            
    return rowoffset

def xlsxExportAdd_tAB(Sheet,Data,rowoffset,coloffset,IName,UName,RName,FName,REName,ALabels,BLabels):
    """
    This function exports a 3D array with aspects time, A, and B to a given excel sheet.
    Same as ExcelExportAdd_tAB but this function is for xlsx files with openpyxl.
    The t dimension is exported in one row, the A and B dimensions as several rows.
    Each row starts with IName (indicator), UName (unit), RName (region), 
    FName (figure where data are used), REName (Resource efficiency scenario), 
    and then come the values for the dimensions A and B and from coloffset onwards, the time dimension.
    Function is meant to be used multiple times, so a rowoffset is given, incremented, and returned for the next run.
    """
    for m in range(0,len(ALabels)):
        for n in range(0,len(BLabels)):
            Sheet.cell(row=rowoffset, column=1).value = IName
            Sheet.cell(row=rowoffset, column=2).value = UName
            Sheet.cell(row=rowoffset, column=3).value = RName
            Sheet.cell(row=rowoffset, column=4).value = FName
            Sheet.cell(row=rowoffset, column=5).value = REName
            Sheet.cell(row=rowoffset, column=6).value = ALabels[m]
            Sheet.cell(row=rowoffset, column=7).value = BLabels[n]
            for t in range(0,Data.shape[0]):
                Sheet.cell(row=rowoffset, column=coloffset + t +1).value = Data[t,m,n]
            rowoffset += 1
            
    return rowoffset

def convert_log(file, file_format='html'):
    """
    Converts the log file to a given file format

    :param file: The filename and path
    :param file_format: The desired format
    """
    output_filename = os.path.splitext(file)[0] + '.' + file_format
    output = pypandoc.convert_file(file, file_format, outputfile=output_filename)
    assert output == ""

def check_dataset(path,PL_Names,PL_Version,PL_SubFolder,Mylog):
    """
    Checks that every parameter in Pl_Names with the corrsponding version PL_Versions is in the folder given by path, or subfolder given by PL_SubFolder

    :param path: Dataset folder
    :param PL_Names: List of parameters names
    :param PL_versions: List of parameters versions
    :param PL_SubFolder: List of data subfolder names
    :param Mylog: log file

    """
    for m in range(len(PL_Names)):
        if PL_Names[m]+'_'+PL_Version[m]+'.xlsx' not in os.listdir(path):
            if PL_Names[m]+'_'+PL_Version[m]+'.xlsx' not in os.listdir(os.path.join(path, PL_SubFolder[m])):
                Mylog.error(PL_Names[m]+'_'+PL_Version[m]+'.xlsx not in the dataset.')
    
    
# The End

